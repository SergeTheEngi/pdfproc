#!/usr/bin/env python
# coding: utf-8

# # This is not the whole notebook, just a quick fix.

# In[20]:


import time

import pymupdf
import re
import copy
import openpyxl
from openpyxl import Workbook
import pdfproc
from pdfproc.as_dict import \
    break_lines, \
    find_line, \
    get_columns, \
    normalize_data, \
    unwrap_sublists_recursive
import pdfproc.as_lines

re_id = '[0-9\\.\\-/A-Z]+'
re_separator = f"\\*+ ?{re_id} ?\\*+"
re_page_end = '\\*+'
# New values
re_hs = 'TAX MAP PARCEL'

ext = pdfproc.as_dict.Extractor(
    re_id,
    re_separator,
    re_page_end
)

# Sources
src_northcastle = pymupdf.open('pdfproc/testing_data:2024FA_Northcastle.pdf')
values_northcastle = {
    'zoning':{},
    'district':{},
}


# In[21]:


# Extract data
alldata = ext.get_data([
    {'source':src_northcastle,'name':'northcastle'},
])

data_northcastle = copy.deepcopy(alldata['northcastle'])

del alldata


# In[26]:


# Get zoning
re_acct = 'ACCT: [A-Z0-9]+'

def find_my_acct(data,key,verbose=False):
    for line in data[key][0]:
        acct = None
        acct = re.search(re_acct,line)
        if acct != None:
            acct = re.split('  +',line)
            for piece in acct:
                if 'ACCT' in piece:
                    out = re.sub('ACCT: ','',piece)
                    return out
    if verbose: print(f"WARN: No account in {key}")
    return key

failed = []
for key in data_northcastle:
    out = find_my_acct(data_northcastle,key)
    if out == key:
        failed.append(out)
        values_northcastle['zoning'][key] = ''
    else:
        values_northcastle['zoning'][key] = out

print(f"Failed to get zoning for:\n{failed}")
print(len(values_northcastle['zoning']),len(data_northcastle))


# In[27]:


for key in values_northcastle['zoning']:
    if len(values_northcastle['zoning'][key]) > 6:
        print(values_northcastle['zoning'][key])


# In[28]:


# Get school districts
re_scl = r'(\w+ )+CENTRAL|CENTRAL (\w+ )+#[0-9]+'

for key in data_northcastle:
    for line in data_northcastle[key][0]:
        scl = None
        line_edit = re.split('  +',line)
        line_edit = '    '.join(line_edit[1:])
        scl = re.search(re_scl,line_edit)
        if scl != None:
            values_northcastle['district'][key] = scl.group()
            break

print(len(values_northcastle['district']),len(data_northcastle))


# ### Write values to the workbook

# In[29]:


# Load workbook
wb = openpyxl.load_workbook(filename='Town of North castle 2024 (Needs Schools and Zoning).xlsx')
ws = wb['2024']


# In[32]:


# Load ids and other data
headers = ws[1]
for item in headers:
    print(item,item.value)
    
ids = ws['A']


# In[33]:


# Write school districs and zoning to new columns
# Add new columns
ws['J1'] = 'ZONING'
ws['K1'] = 'SCHOOL DISTRICT'

# Write zoning values
for key in values_northcastle['zoning']:
    write = False
    for cell in ws['A']:
        if cell.value == key:
            ws[f"J{cell.row}"] = values_northcastle['zoning'][key]
            write=True
            break
    assert write == True
    
# Write district values
for key in values_northcastle['district']:
    write = False
    for cell in ws['A']:
        if cell.value == key:
            ws[f"K{cell.row}"] = values_northcastle['district'][key]
            write=True
            break
    assert write == True

wb.save('Northcastle.xlsx')

