#!/usr/bin/env python
# coding: utf-8

# # Init

# In[1]:


# Monitoring
import time

# PDF extraction engines
import pymupdf
import poppler

# Utilities
import re
import copy
from openpyxl import Workbook

# pdfproc
import pdfproc.as_dict
from pdfproc.as_dict import \
    break_lines, \
    find_line, \
    get_columns, \
    normalize_data, \
    unwrap_sublists_recursive
import pdfproc.as_lines

sources = {}
sources['bronxville'] = pymupdf.open('pdfproc/testing_data:2024FA_Bronxville.pdf')
sources['cornwall'] = pymupdf.open('pdfproc/testing_data:2024FA_Cornwall.pdf')
sources['scarsdale'] = pymupdf.open('pdfproc/testing_data:2024FA_Scarsdale.pdf')
sources['harrison'] = pymupdf.open('pdfproc/testing_data:2024FA_Harrison.pdf')
sources['greenburgh'] = poppler.load_from_file('pdfproc/testing_data:2024FA_Greenburgh.pdf')
sources['newcastle'] = pymupdf.open('pdfproc/testing_data:2024FA_Newcastle.pdf')
sources['mamaroneck'] = poppler.load_from_file('pdfproc/testing_data:2024FA_Mamaroneck.pdf')
sources['bedford'] = poppler.load_from_file('pdfproc/testing_data:2024FA_Bedford.pdf')
sources['mtpleasant'] = poppler.load_from_file('Mount Pleasant.pdf')

re_id = '[0-9\\.\\-/A-Z]+'
re_separator = f"\\*+ ?{re_id} ?\\*+"
re_page_end = '\\*+'
re_hs = 'TAX MAP PARCEL'

ext = pdfproc.as_dict.Extractor(
    re_id,
    re_separator,
    re_page_end
)

lol = pdfproc.as_lines.Extractor(
    re_id=re_id,
    re_separator=re_separator,
    re_hs=re_hs,
)


# # Extraction
# 
# ## Split by entry separators
# ### Get headers
# 
# By block and line number, and assemble it into a new list of the same shape.

# In[2]:


# Test header extractor
bronxville_page = sources['bronxville'].load_page(1)
bronxville_text = bronxville_page.get_text('dict')
cornwall_page = sources['cornwall'].load_page(0)
cornwall_text = cornwall_page.get_text('dict')

testset = [
        (bronxville_text,{
            'get_header':((5,2),(7,0)),
            'assemble_header':[
                [
                    'TAX MAP PARCEL NUMBER        PROPERTY LOCATION & CLASS  ASSESSMENT  EXEMPTION CODE------------------COUNTY--------TOWN------SCHOOL ',
                    'CURRENT OWNERS NAME ',
                    '       SCHOOL DISTRICT  ',
                    '     LAND      TAX DESCRIPTION ',
                    ' ',
                    '  TAXABLE VALUE '
                ],
                [
                    'CURRENT OWNERS ADDRESS        PARCEL SIZE/GRID COORD     TOTAL      SPECIAL DISTRICTS ',
                    ' ',
                    ' ',
                    ' ',
                    ' ACCOUNT NO. '
                ],
            ]
        }),
        (cornwall_text,{
            'get_header':((1,0),(1,3)),
            'assemble_header':[
                [
                    'TAX MAP PARCEL NUMBER          PROPERTY LOCATION & CLASS  ASSESSMENT  EXEMPTION CODE-----VILLAGE------COUNTY--------TOWN------SCHOOL',
                    'CURRENT OWNERS NAME            SCHOOL DISTRICT               LAND      TAX DESCRIPTION            TAXABLE VALUE',
                    'CURRENT OWNERS ADDRESS         PARCEL SIZE/GRID COORD       TOTAL      SPECIAL DISTRICTS                                 ACCOUNT NO.'
                ]
            ]
        })
]

for test,result in testset:
    header_start,header_end = ext.get_header(test)
    assert (header_start,header_end) == result['get_header']
    
    header = ext.assemble_header(test,header_start,header_end)
    assert header == result['assemble_header']


# ### Create entries
# 
# Detect by separators

# In[3]:


# Extract data (list of lines)
import multiprocessing as mp

targets = ['greenburgh','mamaroneck','bedford','mtpleasant']
data = {}

def worker(q,source,name):
    q.put((name,lol.get_data(source)))

with mp.Manager() as m:
    q = m.Queue()
    processes = []

    for t in targets:
        a = [q,sources[t],t]
        p = mp.Process(target = worker, args=a)
        p.start()
        processes.append(p)

    for p in processes:
        p.join()

    while not q.empty():
        t,result = q.get()
        data[t] = result

for line in data['mamaroneck']['1-14-213']:
    if type(line) == str and bool(re.fullmatch('\\*+',line)):
        raise ValueError("Algorithm misses the page end delimiter!")


# In[4]:


# Extract data (pymupdf dictionary)
alldata = ext.get_data([
    {'source':sources['bronxville'],'name':'bronxville'},
    {'source':sources['cornwall'],'name':'cornwall'},
    {'source':sources['scarsdale'],'name':'scarsdale'},
    {
        'source':sources['harrison'],
        'name':'harrison',
        'columns':['TAX MAP','PROPERTY LOCATION','EXEMPTION','TAXABLE VALUE'],
        'strip_lines':False
    },{
        'source':sources['newcastle'],
        'name':'newcastle',
        'columns':['TAX MAP','PROPERTY LOCATION','EXEMPTION','TAXABLE VALUE'],
        'strip_lines':False
    }
])

data_bronxville = copy.deepcopy(alldata['bronxville'])
data_cornwall = copy.deepcopy(alldata['cornwall'])
data_scarsdale = copy.deepcopy(alldata['scarsdale'])
data_harrison = copy.deepcopy(alldata['harrison'])
data_newcastle = copy.deepcopy(alldata['newcastle'])

del alldata


# In[5]:


# Shape data
# Shape cornwall
for key in data_cornwall:
    data_cornwall[key] = unwrap_sublists_recursive(data_cornwall[key])
    data_cornwall[key] = break_lines(data_cornwall[key])

# Shape scarsdale
for key in data_scarsdale:
    data_scarsdale[key] = unwrap_sublists_recursive(data_scarsdale[key])
    data_scarsdale[key] = break_lines(data_scarsdale[key])


# ## Extraction functions & tests

# ### Get owner names

# In[6]:


# Test bronxville
testset_bronxville = [
    ('18./1/2',['Coffey John', 'Coffey Anne', 'Ameriprise Financial-D.Amoruso']),
    ('14./3/4.B',['Hyde Lindsay', 'Hyde Arthur D IV']),
    ('11./5/1.-212',['Nagle,Arthur J, Irrevocable Tr', 'Nagle Christopher P', 'Christopher Nagle']),
    ('4./5/11',['Nibur 132 Parkway Road Bronxvi', 'George Comfort & Sons, Inc.']),
    ('3./3/1.A',['Midland Garden Owners', 'Attn: Barhite & Holzinger']),
    ('1./1/1',['Mercer Robert']),
    ('1./1/10',['Nuguid Dumalag Marie A']),
    ('7.A/3/5',['Copete Andres', 'Copete Margaret M']),
    ('4./1/5',['701 Pondfield LLC', 'Pondfield 17 LLC', 'c/o Houlihan-Parnes Realtors']),
    ('20./2/1.-5L',['Bonanno Rosario']),
    ('15./3/5',['Hannick John D', 'Hannick Elizabeth E']),
    ('6.D/2/10.J',['Wolfe Gregory N', 'Wolfe Elana R']),
    ('1./1/26',['Maianti Echeverrigaray Juan P', 'Darricarrere Nicole']),
    ('1./1/15',['42 Park LLC', 'Mark J. Fonte-Trifont Realty']),
    ('2./3/48',['McLean Heights Realty LLC']),
    ('2./2/17',['Mosbacher Emil','L L C','c/o Mosbacher Properties Group','LLC']),
]

for key,result in testset_bronxville:
    entry = []
    temp_data = normalize_data(data_bronxville[key])
    for block in temp_data:
        if len(block) > 0 and 'PRIOR OWNER' in block[0]: break
        if len(block) > 0:
            if '  ' in block[0]:
                block[0] = block[0].split('  ')[0]
            temp = block[0].replace(' FULL MARKET VALUE','')
            temp = temp.replace('FULL MARKET VALUE','')
            temp = temp.replace(' Bronxville Sch','')
            temp = temp.replace('Bronxville Sch','')
            temp = re.sub(' DEED BOOK.+','',temp)
            temp = re.sub('DEED BOOK.+','',temp)
            entry.append(temp)
    output = ext.get_owner_names(entry,key)
    assert output == result, f"{key}, {result} != {output}"


# In[7]:


# Test cornwall
testset_cornwall = [
    ('101-1-1',['Nguyen Lap','Fowlie Greta']),
    ('102-4-5',['Igo Ellen M']),
]

for key,result in testset_cornwall:
    entry = []
    for line in data_cornwall[key]:
        if 'FULL MARKET VALUE' in line[0] or \
            'DEED BOOK' in line[0] or \
            'EAST-' in line[0] or \
            'ACRES' in line[0] or \
            'add to' in line[0] or \
            'add map' in line[0] or \
            'SCHOOL' in line[0] or \
            'WS SHORE DR' in line[0] or \
            'BLOOMING GROVE' in line[0] or \
            'MAP' in line[0]: break
        elif ('FRNT' in line[0] and 'DPTH' in line[1]):
            if line[0][-4:] == 'FRNT': pass
            else: break
        else:
            patterns = ['FD[0-9]{3}','RG[0-9]{3}']
            temp = None
            append_test = True
            for pattern in patterns:
                temp = re.search(pattern,line[0])
                if temp: append_test = False
            if not append_test: break
            entry.append(line[0])
    output = ext.get_owner_names(entry,key)
    assert output == result, f"{key}, {result} != {output}"


# In[8]:


# Test scarsdale
testset_scarsdale = [
    ('01.05.12',['DALLAL STEVEN','DALLAL TERESA']),
    ('08.19.32',['MCHENRY BRENDAN']),
    ('19.01.40',['MITCHELL ROBERT A','MITCHELL DANA E']),
    ('01.02.20A',['HOFF BARTHELSON','MUSIC SCHOOL','% KEN COLE EXEC DIR']),
    ('11.05.9',['KROHNENGOLD STUART']),
    ('05.02.64',['HITCHCOCK PRESBYTERIAN','CHURCH']),
    ('06.06.4',['LEVINE ERIC B','SIMON TERRI E']),
    ('05.03.72',['NG JOSEPH']),
    ('13.04.52',['NEIRA ALBERT G']),
    ('09.11.50.52',['SILBERBERG HOWARD']),
    ('06.08.18',['LANDAUER BRIAN','LANDAUER MADELYN']),
    ('02.05.9',['SCARSDALE CHATEAUX OWNERS','% BARHITE AND HOLZINGER']),
    ('20.01.7',['636 EXPRESS LLC','SUITE 1191']),
    ('21.01.24',['8 STONEWALL LANE LLC']),
    ('21.01.42',['LEVITT BARRIE']),
    ('10.31.10',['VILLAGE OF SCARSDALE','EAST SIDE CLARENCE RD','TREASURER']),
]

for key,result in testset_scarsdale:
    entry = []
    for line in data_scarsdale[key]:
        if 'FULL MKT VAL' in line[0] or \
            'DEED BK' in line[0] or \
            'EAST-' in line[0] or \
            'ACREAGE' in line[0] or \
            'add to' in line[0] or \
            'add map' in line[0] or \
            'MAP' in line[0] or \
            'TAXABLE' in line[0] or \
            'CONTIGUOUS PARCEL' in line[0] or \
            bool(re.search('[A-Z]{2}[0-9]{3}',line[0])) or \
            bool(re.search('@ [0-9]',line[0])) or \
            bool(re.fullmatch('BANK [0-9]{2,3}',line[0])):
            break
        elif ('FRNT' in line[0] and 'DPTH' in line[0]) or \
            ('FRNT' in line[0] and 'DPTH' in line[1]):
            if line[0][-4:] == 'FRNT': pass
            else: break
        else:
            patterns = [
                ('FD[0-9]{3}',' '.join(line)),
                ('RG[0-9]{3}',' '.join(line)),
                ('[0-9\\,]{4,}+ ?EX',' '.join(line[:2]))
            ]
            temp = None
            append_test = True
            for pattern,text in patterns:
                temp = re.search(pattern,text)
                if temp:
                    append_test = False
            if not append_test:
                break
            else:
                newline = None
                for ln,subline in enumerate(line):
                    if 'SCARSDALE CENTRAL' in subline:
                        if ln > 1:
                            newline = [' '.join(line[:ln])]
                            newline.extend(line[ln:])
                            entry.append(newline[0])
                        else: break
                if not newline:
                    entry.append(line[0])
    #print(entry)
    output = ext.get_owner_names(entry,key)
    assert output == result, f"{key}, {result} != {output}"


# In[9]:


# Test harrison
testset_harrison = [
    ('0011.-1',['RYEWOOD FARMS HOMEOWNERS','ASSOCIATION INC','WESTFAIR PROP MANAG INC','STE 330']),
    ('0856.-17',['MORGADO, RICHARD A']),
    ('0545.-46',['FRANCK, PEGGY MILLER'])
]

for key,result in testset_harrison:
    entry= []
    for line in data_harrison[key]['data'][0][1:]:
        if key in line: entry.append(key)
        else:
            temp = line[data_harrison[key]['columns']['TAX MAP']:data_harrison[key]['columns']['PROPERTY LOCATION']].strip()
            entry.append(temp)
    output = ext.get_owner_names(entry,key)
    assert output == result, f"{key}, {result} != {output}"


# In[10]:


# Test newcastle
testset_newcastle = [
    ('59.13-2-1',['MORGAN TRUST DECLARATION','DUBODEL ALEXANDRA']),
    ('92.20-2-1.-227',['CONYERS VINCENT ERWIN']),
    ('108.5-1-8',['TOWN OF NEW CASTLE']),
    ('101.8-2-9',['BAYSWATER HAMMON RIDGE', 'C/O BAYSWATER DEVELOPMENT']),
    ('71.13-2-29',['SUECHTING RUTH WEIDT', 'IRREVOCABLE TRUST']),
    ('101.8-2-6',['HAMMOND RIDGE HOME-', 'OWNERS ASSOCIATION', 'C/O KATONAH MGT SERVICES'])
]

for key,result in testset_newcastle:
    entry= []
    for line in data_newcastle[key]['data'][0][1:]:
        if key in line: entry.append(key)
        else:
            temp = line[data_newcastle[key]['columns']['TAX MAP']:data_newcastle[key]['columns']['PROPERTY LOCATION']].strip()
            entry.append(temp)
    moved = False
    for n,item in enumerate(entry):
        if bool(re.search('FLOOR [0-9]{2}|[0-9]{1,2}TH FLOOR|PO BOX [0-9]{,4}|#[0-9]{5}|SUITE [0-9A-Z]{,3}',item)):
            if moved:
                moved = False
            else:
                entry[n+1] = ', '.join([item,entry[n+1]])
                entry[n] = ''
                moved = True
    output = ext.get_owner_names(entry,key)
    assert output == result, f"{key}, {result} != {output}"


# In[11]:


# Test greenburgh

testset_greenburgh = [
    ('6.10-1-10.1',['STRONGIN, MEREDITH','STRONGIN, JONATHAN']),
    ('7.280-125-13',['MILLER, KARL M','PARADA, CARLOS']),
    ('8.540-375-12',['DEPIETTO, DAVID']),
    ('6.100-92-1',['PHETSOMPHOU, TERRY']),
    ('7.777-5021-5',['CABLEVISION SYSTEM LIGHTPATH']),
    ('626.89-9999-110.700.2883',['CONSOLIDATED EDISON CO', 'C/O: STEPHANIE J. MERRITT']),
    ('6.80-66-3',['MANSURY, MAWAR U']),
    ('1.100-65-79',['COUNTY OF WESTCHESTER']),
    ('4.100-97-2',['VILLAGE OF HASTINGS ON HUDSON']),
    ('3.60-21-7',['MAYHAWK, JOI']),
    ('5.20-16-13',['TARRY-ELM ASSOCIATES LLC']),
    ('1.271-138-1.232',['BROADWAY ON HUDSON ESTATES LLC']),
    ('2.90-44-21.16',['TOLL NORTHEAST V CORP.']),
]

#print(data['greenburgh'].keys())
#print(data['greenburgh']['6.10-1-10.1'])
for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'DEED BK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK CODE' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('BANK [0-9]{2,3}',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    for ln,line in enumerate(entry):
        entry[ln] = line[0:30]
    try:
        output = ext.get_owner_names(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['greenburgh'][key]: print([line])
        print(entry)
        raise


# In[12]:


# Test mamaroneck
testset_mamaroneck = [
    ('6-1-1',['Moriwaki Yoshizo','Moriwaki Dawn Crista']),
    ('1-14-213',['Del Sole Robert','Del Sole Janet']),
    ('4-5-264',['Borstein David A','Borstein Jessica']),
    ('6-2-96',['Village Of Larchmont','Municipal Building']),
    ('7-14-194',['Breen Adam','Breen Jill']),
]

#print(data['greenburgh'].keys())
#print(data['greenburgh']['6.10-1-10.1'])
for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK CODE' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('BANK [0-9]{2,3}',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    for ln,line in enumerate(entry):
        entry[ln] = line[0:30]
    try:
        output = ext.get_owner_names(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mamaroneck'][key]: print([line])
        print(entry)
        raise


# In[13]:


# Test bedford
testset_bedford = [
    ('38.20-1-1',['Mckenzie Michael']),
    ('60.11-3-43',['Bancroft Robert W']),
    ('84.13-1-3',['Chace Helen Clay','Chace Minturn V Trust','c/o H Chace, M Baker, et., all','as trustees']),
]

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:31])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    try:
        output = ext.get_owner_names(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['bedford'][key]: print([line])
        print(entry)
        raise


# In[14]:


# Test mtpleasant
testset_mtpleasant = [
    ('110.11-1-4',['GUPTA AKSHAY','YANISKO ANNA']),
    ('106.13-1-29',['ANSARI-EZABADI AMIR','MOAZEN HALEH TBTE']),
    ('112.13-3-61',['GAMBINO LISA']),
    ('106.10-1-1.1',['PNMPV REALTY LLC']),
    ('106.5-2-73',['WESTCHESTER COUNTY']),
]

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:26])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                bool(re.search('EAST [0-9]{6}',entry_line[0])),
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                bool(re.fullmatch('[0-9.]{4} [A-Z]{2} [A-Z]{1}',entry_line[0])),
                bool(re.fullmatch('[0-9]-[0-9A-Z]{2,3}-[0-9]{,2}',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    try:
        output = ext.get_owner_names(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mtpleasant'][key]: print([line])
        print(entry)
        raise


# ### Get owner address

# In[15]:


# Test bronxville
testset_bronxville = [
    ('14./3/4.B','5 Edgehill Close, Bronxville, NY 10708'),
    ('2./2/17','18 E 48th Street 19th Floor, New York, NY 10017'),
    ('11./5/17','118-35 Queens Blvd.,1710, Forest Hills, NY 11375'),
    ('11./5/1.-212','906 Mill Creek Dr, Palm Beach Gardens, FL 33410'),
    ('3./3/1.A','77 Pondfield Road, Bronxville, NY 10708'),
    ('20./3/2','33 Sagamore Road, Bronxville, NY 10708'),
    ('1./1/5','60 Parkway Road, Bronxville, NY 10708'),
    ('1./1/10','50 Parkway Road, Bronxville, NY 10708'),
    ('1./1/15','1955 Central Park Ave, Yonkers, NY 10710'),
    ('1./1/15.A','36-38 Parkway Road, Bronxville, NY 10708'),
]

for key,result in testset_bronxville:
    entry = []
    data_temp = normalize_data(data_bronxville[key])
    for block in data_temp:
        if len(block) > 0:
            if 'PRIOR OWNER' in block[0]: break
            company = re.search('(l ?l ?c)|(L ?L ?C)',block[0])
            if company:
                block.pop(0)
                company = None
        else: continue
        if len(block) > 0: entry.append(block[0])
    output = ext.get_owner_address(entry)
    assert output == result, f"{key}, {result} != {output}"


# In[16]:


# Test cornwall
testset_cornwall = [
    ('101-1-1','303 Shore Rd, Cornwall-On-Hudson, NY 12520'),
    ('25-1-3','31 Cedar Ln, Cornwall, NY 12518'),
    ('4-2-8.2','42 Robert Rd, Cornwall, NY 12518'),
    ('116-1-4','184 Mountain Rd, Cornwall-On-Hudson, NY 12520'),
    ('118-1-6.2','25 Boulevard, Cornwall, NY 12518'),
    ('101-1-13.12','16 Idlewild Park Dr, Cornwall-On-Hudson, NY 12520'),
    ('103-1-3.4','378 Hudson St, Cornwall-On-Hudson, NY 12520'),
    #('102-9-24','10-63 Jackson Ave Unit Pha, Long Island City, NY 111001'),
    ('17-6-10','8 Ferguson Rd, Cornwall, NY 12518'),
    ('105-9-17.3','325 Hudson St, Cornwall-On-Hudson, NY 12520'),
    ('1-1-2.22','25 Shore Dr, New Windsor, NY 12553'),
    ('1-1-64.2','1350 Broadway Ste 201, New York, NY 10018'),
    ('1-1-177','18 Idlewild Park Dr, Cornwall on Hudson, NY 12520'),
    ('1-2-19','255 Orrs Mills Rd, Salisbury Mills, NY 12577'),
]

for key,result in testset_cornwall:
    entry = []
    for line in data_cornwall[key]:
        if 'FULL MARKET VALUE' in line[0] or \
            'DEED BOOK' in line[0] or \
            'EAST-' in line[0] or \
            'ACRES' in line[0] or \
            'add to' in line[0] or \
            'add map' in line[0] or \
            'SCHOOL' in line[0] or \
            'WS SHORE DR' in line[0] or \
            'BLOOMING GROVE' in line[0] or \
            'MAP' in line[0]: break
        elif ('FRNT' in line[0] and 'DPTH' in line[1]):
            if line[0][-4:] == 'FRNT': pass
            else: break
        else:
            patterns = ['FD[0-9]{3}','RG[0-9]{3}']
            temp = None
            append_test = True
            for pattern in patterns:
                temp = re.search(pattern,line[0])
                if temp: append_test = False
            if not append_test: break
            entry.append(line[0])
    output = ext.get_owner_address(entry)
    assert output == result, f"{key}, {result} != {output}"


# In[17]:


# Test scarsdale
testset_scarsdale = [
    ('01.05.12','5 SCHOOL LA, SCARSDALE NY 10583'),
    ('08.19.32','113 BRAMBACH RD, SCARSDALE NY 10583'),
    ('19.01.40','18 CORALYN RD, SCARSDALE NY 10583'),
    ('01.02.20A','25 SCHOOL LA, SCARSDALE NY 10583'),
    ('11.05.9','15 HAMILTON RD, SCARSDALE NY 10583'),
    ('05.02.64','6 GREENACRES AVE, SCARSDALE NY 10583'),
    ('09.11.50.52','37 SPRAGUE RD, SCARSDALE NY 10583'),
    ('06.08.18','1 BERKELEY RD, SCARSDALE NY 10583'),
    ('01.01.4','10 SCHOOL LA, SCARSDALE NY 10583'),
    ('02.05.9','77 PONDFIELD RD, BRONXVILLE NY 10708'),
    ('20.01.7','455 TARRYTOWN RD, WHITE PLAINS NY 10607'),
    ('21.01.24','16 STONEWALL LA, MAMARONECK NY 10538'),
    ('21.01.42','16 STONEWALL LA, MAMARONECK NY 10543'),
    ('10.31.10','1001 POST RD, SCARSDALE NY 10583'),
]

for key,result in testset_scarsdale:
    entry = []
    for line in data_scarsdale[key]:
        if 'FULL MKT VAL' in line[0] or \
            'DEED BK' in line[0] or \
            'EAST-' in line[0] or \
            'ACREAGE' in line[0] or \
            'add to' in line[0] or \
            'add map' in line[0] or \
            'MAP' in line[0] or \
            'TAXABLE' in line[0] or \
            'CONTIGUOUS PARCEL' in line[0] or \
            bool(re.search('[A-Z]{2}[0-9]{3}',line[0])) or \
            bool(re.search('@ [0-9]',line[0])) or \
            bool(re.fullmatch('BANK [0-9]{2,3}',line[0])):
                break
        elif ('FRNT' in line[0] and 'DPTH' in line[0]) or ('FRNT' in line[0] and 'DPTH' in line[1]):
            if line[0][-4:] == 'FRNT': pass
            else:
                break
        else:
            patterns = [
                ('FD[0-9]{3}',' '.join(line)),
                ('RG[0-9]{3}',' '.join(line)),
                ('[0-9\\,]{4,}+ ?EX',' '.join(line[:2]))
            ]
            temp = None
            append_test = True
            for pattern,text in patterns:
                temp = re.search(pattern,text)
                if temp:
                    if key == '10.31.10': print(temp.group())
                    append_test = False
            if not append_test:
                break
            else:
                entry.append(line[0])
    output = ext.get_owner_address(entry)
    assert output == result, f"{key}, {result} != {output}"


# In[18]:


# Test harrison
testset_harrison = [
    ('0011.-1','56 LAFAYETTE AVE, WHITE PLAINS NY 10603'),
    ('0824.-19','146 FREMONT ST, HARRISON NY 10528'),
    ('0892.-32','105 CORPORATE PARK DR-APT, WEST HARRISON NY 10604')
]

for key,result in testset_harrison:
    entry= []
    for line in data_harrison[key]['data'][0][1:]:
        start = data_harrison[key]['columns']['TAX MAP']
        end = data_harrison[key]['columns']['PROPERTY LOCATION']
        temp = line[start:end].strip()
        entry.append(temp)
    output = ext.get_owner_address(entry)
    assert output == result, f"{key}, {[result]} != {[output]}"


# In[19]:


# Test newcastle
testset_newcastle = [
    ('999.999-2-28','STATE OF NEW YORK'),
    ('101.15-1-13','148 MARTINE AVE, WHITE PLAINS NY 10601'),
    ('91.12-1-5','171 CAMPFIRE RD, CHAPPAQUA NY 10514'),
    ('91.15-1-22','STATE OF NEW YORK'),
    ('101.8-2-9','PO BOX 7430, PORT ST. LUCIE, FLORIDA 34985')
]

for key,result in testset_newcastle:
    entry= []
    for line in data_newcastle[key]['data'][0][1:]:
        start = data_newcastle[key]['columns']['TAX MAP']
        end = data_newcastle[key]['columns']['PROPERTY LOCATION']
        temp = line[start:end].strip()
        entry.append(temp)
    for n,item in enumerate(entry):
        if bool(re.search('FLOOR [0-9]{2}|[0-9]{1,2}TH FLOOR|PO BOX [0-9]{,4}|#[0-9]{5}|SUITE [0-9A-Z]{,3}',item)):
            if moved:
                moved = False
            else:
                entry[n+1] = ', '.join([item,entry[n+1]])
                entry[n] = ''
                moved = True
    output = ext.get_owner_address(entry,key)
    assert output == result, f"{key}, {[result]} != {[output]}"


# In[20]:


# Test greenburgh
testset_greenburgh = [
    ('6.10-1-10.1','36 CONCORD RD, ARDSLEY NY 10502'),
    ('7.280-125-13','50 FISHER LN, WHITE PLAINS NY 10603'),
    ('8.540-375-12','22 WALBROOKE RD, SCARSDALE NY 10583'),
]
for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'DEED BK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK CODE' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('BANK [0-9]{2,3}',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    for ln,line in enumerate(entry):
        entry[ln] = line[0:30]
    try:
        output = ext.get_owner_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['greenburgh'][key]: print([line])
        print(entry)
        raise


# In[21]:


# Test mamaroneck

testset_mamaroneck = [
    ('6-8-307','17 Hall Ave, Larchmont, NY 10538'),
    ('1-14-314','121 N Chatsworth Ave, Larchmont, NY 10538'),
    ('2-19-311','19 Lansdowne Dr, Larchmont, NY 10538'),
]
for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK CODE' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('BANK [0-9]{2,3}',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    for ln,line in enumerate(entry):
        entry[ln] = line[0:30]
    try:
        output = ext.get_owner_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mamaroneck'][key]: print([line])
        print(entry)
        raise


# In[22]:


# Test bedford
testset_bedford = [
    ('38.20-1-1','171 Goldens Bridge Rd, Katonah, NY 10536'),
    ('60.11-3-43','10 Franklin Ave, Bedford Hills, NY 10507'),
    ('84.13-1-3','20 Bayberry Ln, Bedford Corners, NY 10549'),
]

#print(data['greenburgh'].keys())
#print(data['greenburgh']['6.10-1-10.1'])
for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:31])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    try:
        output = ext.get_owner_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['bedford'][key]: print([line])
        print(entry)
        raise


# In[23]:


# Test mtpleasant
testset_mtpleasant = [
    ('106.9-1-59','579 BEDFORD ROAD, PLEASANTVILLE NY 10570'),
    ('105.8-2-26','18 WOODFIELD ROAD, BRIARCLIFF MANOR NY 10510'),
    ('112.13-3-61','230 BRADY AVE, HAWTHORNE NY 10532'),
]
# Unsolved: ('106.5-2-73',''),

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:26])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                bool(re.search('EAST [0-9]{6}',entry_line[0])),
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                bool(re.fullmatch('[0-9.]{4} [A-Z]{2} [A-Z]{1}',entry_line[0])),
                bool(re.fullmatch('[0-9]-[0-9A-Z]{2,3}-[0-9]{,2}',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    try:
        output = ext.get_owner_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mtpleasant'][key]: print([line])
        print(entry)
        raise


# ### Get property type

# In[24]:


def get_property_type(entry,key):
    try:
        if entry[1][1] == '':
            if entry[1][3] == '': return entry[1][2]
            else: return ' '.join(entry[1][2:4])
        elif entry[1][1] == key or entry[2][0] == key:
            return entry[2][1]
        else: return entry[1][1]
    except Exception as e:
        print(f"WARN: {e} for {key}")
        return entry[2][1]

def run_test(entry,key,result,function):
    output = function(entry,key)
    assert output == result, f"{key}, {result} != {output}"

# Tests
testset_bronxville = [
    ('13./3/1','210 1 Family Res'),
    ('11./3/3.A','311 Res vac land'),
    ('20./2/1.-7K','411 Apartment - CONDO'),
    ('20./2/60.A-8A','210 1 Family Res'),
]

testset_cornwall = [
    ('101-1-1','210 1 Family Res'),
    ('25-1-3','220 2 Family Res'),
    ('39-6-15','250 Estate'),
    ('101-3-1','210 1 Family Res'),
    ('101-1-13.12','210 1 Family Res'),
    ('101-3-3','210 1 Family Res'),
    ('1-1-64.2','311 Res vac land'),
]

testset_scarsdale = [
    ('01.05.12','210 1 FAMILY RES'),
    ('08.19.32','210 1 FAMILY RES'),
    ('19.01.40','210 1 FAMILY RES'),
    ('01.02.20A','612 SCHOOLS'),
    ('11.05.9','210 1 FAMILY RES'),
]

testset_harrison = [
    ('0011.-1','311 RES VACANT LAND'),
    ('0601.-17','210 1 FAMILY RES'),
    ('1002.-27','210 1 FAMILY RES')
]

testset_newcastle = [
    ('59.13-2-1','210 1 FAMILY RES'),
    ('999.999-1-975.1','860 SPECIAL FRANCHISE PR'),
    ('92.14-1-1','210 1 FAMILY RES')
]

for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_property_type)

for key,result in testset_cornwall:
    entry = data_cornwall[key]
    if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    run_test(entry,key,result,get_property_type)

for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key])
    if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    trash = None; trash = re.search('[A-Z]{2}',entry[2][1])
    if trash: del entry[2][1]
    run_test(entry,key,result,get_property_type)

for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)
    if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    trash = None; trash = re.search('[A-Z]{2}',entry[2][1])
    if trash: del entry[2][1]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_property_type)

for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)
    if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    trash = None; trash = re.search('[A-Z]{2}',entry[2][1])
    if trash: del entry[2][1]
    for n,e in enumerate(entry):
        for ln,line in enumerate(e):
            if bool(re.fullmatch('[A-Z]{2}',line)):
                e[ln] = ''
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_property_type)

all_types = []
for entry in data_bronxville:
    all_types.append(get_property_type(data_bronxville[entry],entry))

assert '' in ['', 'test']
assert '' not in all_types


# In[25]:


# Test greenburgh
testset_greenburgh = [
    ('6.10-1-10.1','210 One Family Year-Round Residence'),
    ('7.280-125-13','210 One Family Year-Round Residence'),
    ('8.540-375-12','210 One Family Year-Round Residence'),
]

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            #print(entry_line)
            entry.append(entry_line)
    try:
        output = get_property_type(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['greenburgh'][key]: print([line])
        print(entry)
        raise


# In[26]:


# Test mamaroneck
testset_mamaroneck = [
    ('8-26-61','714 Lite Ind Man'),
    ('1-33-599.65','411 Apartment'),
    ('2-19-341','210 1 Family Res'),
]

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            #print(entry_line)
            entry.append(entry_line)
    try:
        output = get_property_type(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mamaroneck'][key]: print([line])
        print(entry)
        raise


# In[27]:


# Test bedford
testset_bedford = [
    ('60.5-2-17','210 1 Family Res'),
    ('60.13-1-152','210 1 Family Res - CONDO'),
    ('84.12-2-7','210 1 Family Res'),
]

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            #print(entry_line)
            entry.append(entry_line)
    try:
        output = get_property_type(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['bedford'][key]: print([line])
        print(entry)
        raise


# In[28]:


# Test mtpleasant
testset_mtpleasant = [
    ('110.11-1-3','641 HOSPITALS'),
    ('112.6-4-59','210 1 FAMILY RES'),
    ('112.16-1-67','210 1 FAMILY RES'),
]

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            for sn, substring in enumerate(entry_line):
                if substring == key and bool(re.fullmatch('[A-Z]{2}',entry_line[sn+1])):
                    del entry_line[sn+1]
                    break
            #print(entry_line)
            entry.append(entry_line)
    try:
        output = get_property_type(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        for line in data['mtpleasant'][key]: print([line])
        print(entry)
        raise


# ### Get property address

# In[29]:


# Mix of function definition and tests
def get_property_address(entry,key):
    address = [line for line in entry[0][1:] if line != '']
    if key in address:
        print(f"WARN: {key} has no address")
        return None
    if len(address) > 1:
        checks = [
            len([i for i in address if 'INCLUDLOT' in i]) != 1,
            len(address) == 2 and address[0] != address[1],
            '836' not in get_property_type(entry,key)
        ]
        if True in checks:
            raise ValueError(f"entry {entry[1][0]} in {key}: Address len is more than 1")
    if len(address) == 0:
        print(f"WARN: {key} has no address")
        return ''
    else:
        return address[0]

def run_test(entry,key,result,function):
    output = function(entry,key)
    assert output == result, f"{key}, {result} != {output}"

# Tests
testset_bronxville = [
    ('7.H/2/2','26 Courseview Road'),
    ('11./5/7', None),
]

testset_cornwall = [
    ('101-1-1','303 Shore Rd'),
    ('16-11-12','13 Bede Ter'),
    ('30-1-97.1','2 J R Ct'),
    ('101-3-1','40 Pine St'),
    ('107-2-34.5','61 Duncan Ave'),
    ('1-1-62.1','10 Morris Ln'),
    ('1-1-64.2',None),
    ('2-3-10','8 Oak Dr'),
    ('13-1-11','21 Howard St'),
    ('22-1-15.1','22 Willow Ave'),
]

testset_scarsdale = [
    ('01.05.12','5 SCHOOL LA'),
    ('08.19.32','113 BRAMBACH RD'),
    ('19.01.40','18 CORALYN RD'),
    ('01.02.20A','25 SCHOOL LA'),
    ('11.05.9','15 HAMILTON RD'),
    ('02.05.2A.2','CHRISTIE PL'),
]

testset_harrison = [
    ('0011.-1','HARRISON AVE'),
    ('0422.-10','93 WHITE ST'),
    ('0883.-31','207-209 GAINSBORG AVE')
]

testset_newcastle = [
    ('92.14-1-1','14 VALLEY VIEW'),
    ('71.9-1-9.-66','66 BURR FARMS RD'),
    ('999.999-1-975.1','SPECIAL FRANCHISE'),
    ('91.11-2-53','')
]

for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_property_address)

for key,result in testset_cornwall:
    entry = copy.deepcopy(data_cornwall[key][1:])
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
        if len(entry[0]) > 2:
            entry[0][1] = ' '.join(entry[0][1:])
            entry[0] = entry[0][0:2]
    else: entry[0].append(entry[0][0])
    entry[0][1] = re.sub(' [0-9 A-Z]{10,}+','',entry[0][1])
    run_test(entry,key,result,get_property_address)

for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key][1:])
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        entry[0][1] = ' '.join(entry[0][1:])
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    run_test(entry,key,result,get_property_address)

for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        entry[0][1] = ' '.join(entry[0][1:])
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    run_test(entry,key,result,get_property_address)

for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        for ln,line in enumerate(entry[n]):
            entry[n][ln] = re.sub('(NON-)?HOMESTEAD','',line)
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        entry[0][1] = ' '.join(entry[0][1:])
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    run_test(entry,key,result,get_property_address)


# In[30]:


# Test greenburgh
testset_greenburgh = [
    ('6.10-1-10.1','36 CONCORD RD'),
    ('7.280-125-13','50 FISHER LN'),
    ('8.540-375-12','22 WALBROOKE RD'),
    ('6.50-31-17','0 BEACON HILL RD (OFF)'),
    ('6.100-92-10','0 ALMENA AVE'),
]

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    del entry[0]
    nentry = []
    for line in entry[0]:
        checks = [
            not bool(re.fullmatch('[0-9]+',line)),
            not bool(re.search('ACCT:',line)),
        ]
        if False not in checks:
            nentry.append(line)
    entry[0] = nentry
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        while len(entry[0]) > 2:
            if 'ACCT' not in entry[0][2]:
                entry[0][1] = entry[0][1] + ' ' + entry[0][2]
                del entry[0][2]
            else:
                break
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    try:
        output = get_property_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        raise


# In[31]:


# Test mamaroneck
testset_mamaroneck = [
    ('8-26-208','714-16 Waverly Ave'),
    ('2-17-835','104 W Garden Rd'),
    ('2-25-1.2','Weaver St'),
]

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    del entry[0]
    nentry = []
    for line in entry[0]:
        checks = [
            not bool(re.fullmatch('[0-9]+',line)),
            not bool(re.search('ACCT:',line)),
        ]
        if False not in checks:
            nentry.append(line)
    entry[0] = nentry
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        while len(entry[0]) > 2:
            if 'ACCT' not in entry[0][2]:
                entry[0][1] = entry[0][1] + ' ' + entry[0][2]
                del entry[0][2]
            else:
                break
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    try:
        output = get_property_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mamaroneck'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        raise


# In[32]:


# Test bedford
testset_bedford = [
    ('60.5-2-17','67 Lily Pond Ln'),
    ('60.13-1-152','14 Lake Marie Ln'),
    ('84.12-2-7','189 Pound Ridge Rd'),
]

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    del entry[0]
    nentry = []
    for line in entry[0]:
        checks = [
            not bool(re.fullmatch('[0-9]+',line)),
            not bool(re.search('ACCT:',line)),
        ]
        if False not in checks:
            nentry.append(line)
    entry[0] = nentry
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        while len(entry[0]) > 2:
            if 'ACCT' not in entry[0][2]:
                entry[0][1] = entry[0][1] + ' ' + entry[0][2]
                del entry[0][2]
            else:
                break
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    try:
        output = get_property_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['bedford'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        raise


# In[33]:


# Test mtpleasant
testset_mtpleasant = [
    ('106.7-1-66','21 WILLOW ST'),
    ('106.13-3-26','PARK ST'),
    ('106.20-3-42.12','7 COTTAGE GROVE'),
]

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    del entry[0]
    nentry = []
    for line in entry[0]:
        checks = [
            not bool(re.fullmatch('[0-9]+',line)),
            not bool(re.search('ACCT:',line)),
        ]
        if False not in checks:
            nentry.append(line)
    entry[0] = nentry
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        while len(entry[0]) > 2:
            if 'ACCT' not in entry[0][2]:
                entry[0][1] = entry[0][1] + ' ' + entry[0][2]
                del entry[0][2]
            else:
                break
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    try:
        output = get_property_address(entry,key)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mtpleasant'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        raise


# ### Get zoning

# In[34]:


# Tests and the function definition
def get_zoning(entry,pattern):
    lines = [
        ' '.join(entry[1]),
        ' '.join(entry[2])
    ]
    for item in lines:
        zoning = re.search(pattern, item)
        if zoning:
            zoning = zoning.group()
            zoning = re.sub(' +', ' ',zoning)
            zoning = zoning.strip()
            return zoning
        
    #zoning = entry[2][1]
    #if zoning == '': zoning = entry[2][2]
    #if 'Bronxville Sch' in zoning:
    #    return entry[2][1]
    #else:
    #    print(entry[1][0])

def run_test(entry,key,result,function,pattern):
    output = function(entry,pattern)
    assert output == result, f"{key}, {result} != {output}"

# Tests
testset_bronxville = [
    ('1./1/15.A','Bronxville Sch 552403'),
    ('13./4/3', 'Bronxville Sch 552403'),
]

testset_cornwall = [
    ('30-1-97.1','Cornwall Csd 332401'),
    ('24-1-14','Cornwall Csd 332401'),
    ('1-1-64.2','Washingtonville 332002'),
    ('5-3-3','Newburgh Csd 331100'),
]

testset_scarsdale = [
    ('01.05.12','SCARSDALE CENTRAL'),
    ('08.19.32','SCARSDALE CENTRAL'),
    ('19.01.40','SCARSDALE CENTRAL'),
    ('01.02.20A','SCARSDALE CENTRAL'),
    ('11.05.9','SCARSDALE CENTRAL'),
]

testset_harrison = [
    ('0883.-31','HARRISON CENTRAL'),
    ('0103.-13','HARRISON CENTRAL'),
    ('1005.-32','HARRISON CENTRAL')
]

testset_newcastle = [
    ('71.9-1-14','YORKTOWN CENTRAL'),
    ('71.13-3-17','CHAPPAQUA CENTRAL'),
    ('79.15-1-12','OSSINING UNION FREE')
]

for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_zoning,'Bronxville Sch  ?\\d{6} ')

for key,result in testset_cornwall:
    entry = copy.deepcopy(data_cornwall[key][1:])
    #if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    run_test(entry,key,result,get_zoning,'Cornwall Csd  ?\\d{6} |Washingtonville  ?\\d{6} |[A-Za-z]+ Csd  ?\\d{6}')

# Scarsdale tests
for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key][1:])
    #if any(('FULL MARKET VALUE' in i) for i in entry[-1]): entry = entry[:-1]
    run_test(entry,key,result,get_zoning,'SCARSDALE CENTRAL')

for key in data_scarsdale:
    result = None; entry = None
    entry = copy.deepcopy(data_scarsdale[key][1:])
    result = get_zoning(entry,'SCARSDALE CENTRAL')
    assert result != None

# Harrison tests
for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    run_test(entry,key,result,get_zoning,'HARRISON CENTRAL')

for key in data_harrison:
    result = None; entry = None
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    result = get_zoning(entry,'HARRISON CENTRAL')
    assert result != None

# Newcastle tests
for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    run_test(entry,key,result,get_zoning,'(YORKTOWN|CHAPPAQUA|BEDFORD|BYRAM HILLS) CENTRAL|OSSINING UNION FREE|PLEASANTVILLE UFSD')

for key in data_newcastle:
    result = None; entry = None
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    result = get_zoning(entry,'(YORKTOWN|CHAPPAQUA|BEDFORD|BYRAM HILLS) CENTRAL|OSSINING UNION FREE|PLEASANTVILLE UFSD')
    assert result != None, f"get_zoning({key}) == None"

# Old tests
all_zoning = []
for entry in data_bronxville:
    all_zoning.append(get_zoning(data_bronxville[entry],'Bronxville Sch  ?\\d{6} '))
    if get_zoning(data_bronxville[entry],'Bronxville Sch  ?\\d{6} |Washingtonville  ?\\d{6} ') == None:
        print(entry)
        for block in data_bronxville[entry]:
            print(block)
    #print(entry,"\t:",get_zoning(data[entry]))

assert '' in ['', 'test']
assert '' not in all_zoning
assert None not in all_zoning


# In[35]:


# Test greenburgh with the generic function (SCHOOL DISTRICT)
testset_greenburgh = [
    ('6.10-1-10.1','ARDSLEY'),
    ('7.280-125-13','GREENBURGH'),
    ('7.280-125-3','VALHALLA'),
    ('8.540-375-12','EDGEMONT'),
]

re_zoning = 'ARDSLEY|VALHALLA|EDGEMONT|GREENBURGH|DOBBS FERRY|ELMSFORD|HASTINGS|IRVINGTON|TARRYTOWN|POCANTICO'

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry[3],re_zoning)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise
        
for key in data['greenburgh']:
    entry = []
    verbose = False
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_zoning,verbose)
    except:
        print(entry)
        raise


# In[36]:


# Test greenburgh with the generic function (ACCT)
testset_greenburgh = [
    ('6.10-1-10.1','ACCT: 6167550'),
    ('7.280-125-13','ACCT: 7255425'),
    ('8.540-375-12','ACCT: 8531402'),
]

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,'ACCT: [0-9]{6,7}')
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

for key in data['greenburgh']:
    entry = []
    verbose = False
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,'ACCT: [0-9]{6,7}',verbose)
    except:
        print(entry)
        raise


# In[37]:


# Test mamaroneck with the generic function (SCHOOL DISTRICT)
testset_mamaroneck = [
    ('2-28-78','Mamaroneck 553201'),
    ('8-29-226','Mamaroneck 553201'),
    ('6-4-315','Mamaroneck 553201'),
]

re_zoning = '(Mamaroneck|Scarsdale) [0-9]{6,7}'

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_zoning)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mamaroneck'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

for key in data['mamaroneck']:
    entry = []
    verbose = False
    for line in data['mamaroneck'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_zoning,verbose)
    except:
        print(entry)
        raise


# In[38]:


# Test bedford with the generic function (SCHOOL DISTRICT)
testset_bedford = [
    ('49.11-2-13','Katonah Lewisbo 552001'),
    ('72.5-2-2','Bedford Central 552002'),
    ('84.10-2-8','Bedford Central 552002'),
]

re_zoning = '(Bedford Central|Katonah Lewisbo|Byram Hills Cnt) [0-9]{6,7}'

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_zoning)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['bedford'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

for key in data['bedford']:
    entry = []
    verbose = False
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_zoning,verbose)
    except:
        print(entry)
        raise


# In[39]:


# Test mtpleasant with the generic function (SCHOOL DISTRICT)
testset_mtpleasant = [
    ('104.15-1-3','NORTH TARRYTOWN'),
    ('106.20-1-67','MOUNT PLEASANT'),
    ('117.6-1-8','VALHALLA'),
]

re_zoning = '(NORTH TARRYTOWN|MOUNT PLEASANT|VALHALLA|POCANTICO HILLS|PLEASANTVILLE|BYRAM HILLS CENTRL|BRIARCLIFF MANOR|CHAPPAQUA CENTRAL|HAWTHORNE KNOLL)'

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_zoning)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mtpleasant'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

for key in data['mtpleasant']:
    entry = []
    verbose = False
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_zoning,verbose)
    except:
        print(entry)
        raise


# ### Get acreage

# In[40]:


def get_acreage(entry,keyword='ACRES'):
    for block in entry:
        temp = ' '.join(block)
        temp = ' '.join(temp.split())
        acres = re.search(f"{keyword} [0-9]+\\.[0-9][0-9]",temp)
        if acres:
            acres = re.sub(f'{keyword} ','',acres.group())
            return float(acres)

# Tests
def run_test(entry,key,result,function,keyword='ACRES'):
    output = function(entry,keyword)
    assert output == result, f"{key}, {result} != {output}"

testset_bronxville = [
    ('12./3/13', 0.30),
    ('1./1/14', 0.05),
    ('7.F/5/2', None),
]

testset_cornwall = [
    ('2-9-12', 0.15),
    ('35-1-21.2', 1.00),
    ('624.089-0000-621.700-1881', None),
    ('29-1-13.221',15.90),
    ('106-3-36',0.94),
]

testset_scarsdale = [
    ('01.05.12',0.40),
    ('08.19.32',0.11),
    ('19.01.40',0.16),
    ('01.02.20A',1.48),
    ('11.05.9',0.18),
    ('08.22.20',None),
]

testset_harrison = [
    ('0011.-1',1.10),
    ('0883.-31',0.11),
    ('0103.-13',0.05),
    ('1005.-32',1.07),
]

testset_newcastle = [
    ('101.15-1-11',3.37),
    ('80.19-2-18',1.00),
    ('93.6-2-12',1.00)
]

for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_acreage)

for key,result in testset_cornwall:
    entry = copy.deepcopy(data_cornwall[key][1:])
    run_test(entry,key,result,get_acreage)

for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key][1:])
    run_test(entry,key,result,get_acreage,'ACREAGE')

for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_acreage,'ACREAGE')

for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_acreage,'ACREAGE')


all_acreage = []
for entry in data_bronxville:
    acr = get_acreage(data_bronxville[entry])
    all_acreage.append(acr)

assert '' in ['', 'test']
assert '' not in all_acreage


# In[41]:


# Test greenburgh with the new generic function
testset_greenburgh = [
    ('6.10-1-10.1',.2999),
    ('7.280-125-13',.2099),
    ('8.540-375-12',.1999),
]

re_acreage = 'ACREAGE [0-9.]{,7}'

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_acreage)
        output = output.split()
        output = float(output[1])
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

for key in data['greenburgh']:
    entry = []
    verbose = False
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_acreage,verbose)
    except:
        print(entry)
        raise


# In[42]:


# Test mamaroneck with the new generic function
testset_mamaroneck = [
    ('2-1-1',0.56),
    ('2-25-1.1./2',9.00),
    ('4-8-154',0.09),
]

re_acreage = 'ACRES ?[0-9.]{,7}'

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_acreage)
        output = output.split()
        output = float(output[1])
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mamaroneck'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

failed = []
for key in data['mamaroneck']:
    entry = []
    verbose = False
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_acreage,verbose)
    except:
        failed.append(key)

print(f"Failed to find acreage in {len(failed)} entries")


# In[43]:


# Test bedford with the new generic function
testset_bedford = [
    ('49.10-3-5',None),
    ('60.14-3-27',0.34),
    ('72.5-1-9',1.03),
]

re_acreage = 'ACRES ?[0-9.]{,7}'

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_acreage)
        output = output.split()
        output = float(output[1])
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Failed to get acreage!")
        print("Data item:")
        for line in data['bedford'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        output = None
        assert output == result, f"{key}, {result} != {output}"

failed = []
for key in data['bedford']:
    entry = []
    verbose = False
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_acreage,verbose)
    except:
        failed.append(key)

print(f"\nFailed to find acreage in {len(failed)} entries")


# In[44]:


# Test mtpleasant with the new generic function
testset_mtpleasant = [
    ('115.15-1-16',0.05),
    ('106.9-4-18',0.55),
    ('122.7-4-7',0.11),
]

re_acreage = '(ACREAGE|ACRES) ?[0-9.]{,7}'

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_acreage)
        output = output.split()
        output = float(output[1])
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Failed to get acreage!")
        print("Data item:")
        for line in data['mtpleasant'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        output = None
        assert output == result, f"{key}, {result} != {output}"

failed = []
for key in data['mtpleasant']:
    entry = []
    verbose = False
    for line in data['mtpleasant'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_acreage,verbose)
    except:
        failed.append(key)

print(f"\nFailed to find acreage in {len(failed)} entries")


# ### Get full market value

# In[45]:


def get_full_market_value(entry,keywords=['FULL MARKET VALUE','VALUE']):
    for block in entry:
        for ln,line in enumerate(block):
            if keywords[0] in line:
                if line == keywords[0]:
                    if block[ln+1] == '':
                        mval = block[ln+2]
                    else:
                        mval = block[ln+1]
                    if type(mval) == str:
                        if re.search('[a-zA-Z ]', mval):
                            mval = mval.split()[0]
                        return float(mval.replace(',',''))
                else:
                    line = line.split()
                    for sn,subline in enumerate(line):
                        if subline == keywords[1]:
                            try: mval = line[sn+1]
                            except Exception as e:
                                if block[ln+1] == '':
                                    mval = block[ln+2]
                                else:
                                    mval = block[ln+1]
                            break
                    if type(mval) == str:
                        if re.search('[a-zA-Z ]', mval):
                            mval = mval.split()[0]
                        return float(mval.replace(',',''))
                                
                        
# Tests
def run_test(entry,key,result,function,keywords=['FULL MARKET VALUE','VALUE']):
    output = function(entry,keywords)
    assert output == result, f"{key}, {result} != {output}"

testset_bronxville = [
    ('11./5/1.-202', 1011081),
    ('2./1/11', 6321803),
]

testset_cornwall = [
    ('102-19-10',424600),
    ('106-3-36',761400),
    ('20-6-12',551600),
    ('624.089-0000-621.700-1881',0)
]

testset_scarsdale = [
    ('01.05.12',2568721),
    ('08.19.32',1067520),
    ('19.01.40',2297037),
    ('01.02.20A',11742727),
    ('11.05.9',1177441),
]

testset_harrison = [
    ('0011.-1',176991),
    ('0883.-31',1150442),
    ('0103.-13',113274),
    ('1005.-32',1765486),
]

testset_newcastle = [
    ('93.6-2-12',672281),
    ('636.798-760-884',142),
    ('999.999-1-547.80',2654477)
]

for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_full_market_value)

for key,result in testset_cornwall:
    entry = copy.deepcopy(data_cornwall[key][1:])
    run_test(entry,key,result,get_full_market_value)

for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key][1:])
    run_test(entry,key,result,get_full_market_value,keywords=['FULL MKT VAL','VAL'])

for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_full_market_value,keywords=['FULL MKT VAL','VAL'])

for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_full_market_value,keywords=['FULL MKT VAL','VAL'])

all_market_values = []
for entry in data_bronxville:
    mval = get_full_market_value(data_bronxville[entry])
    all_market_values.append(mval)
    if mval == None: print(data[entry])
    #try:all_market_values.append(get_full_market_value(data[entry]))
    #except Exception as e:
        #print(entry,":",e)
        #print(entry)

assert '' in ['', 'test']
assert '' not in all_market_values
assert None not in all_market_values


# In[46]:


# Test greenburgh with the new generic function
testset_greenburgh = [
    ('6.10-1-10.1',1487500),
    ('7.280-125-13',619200),
    ('8.540-375-12',979200),
]

re_fmv = 'FULL MKT VAL [0-9,]+'

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

failed = []
for key in data['greenburgh']:
    entry = []
    verbose = False
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        ext.get_generic(entry,re_fmv,verbose)
    except:
        try:
            ext.get_generic(entry,'FULL MKT VAL',verbose)
            failed.append(key)
        except:
            print(entry)
            raise

print(f"Entries without full market value:\n{failed}")


# In[47]:


# Test mamaroneck with the new generic function
testset_mamaroneck = [
    ('7-21-103',1842000),
    ('9-55-105',4494000),
    ('1-34-PO.9',7700),
]

re_fmv = 'FULL MKT VAL [0-9,]+|FULL MARKET VALUE [0-9,]+'

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mamaroneck'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

failed = []
for key in data['mamaroneck']:
    entry = []
    verbose = False
    for line in data['mamaroneck'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
    except:
        failed.append(key)

print(f"Entries without full market value:\n{failed}")


# In[48]:


# Test bedford with the new generic function
testset_bedford = [
    ('49.14-2-30',769643),
    ('73.12-1-2',5178571),
    ('84.16-2-29',2832020),
]

re_fmv = 'FULL MKT VAL [0-9,]+|FULL MARKET VALUE [0-9,]+'

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['bedford'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

failed = []
for key in data['bedford']:
    entry = []
    verbose = False
    for line in data['bedford'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
    except:
        failed.append(key)

print(f"Entries without full market value:\n{failed}")


# In[49]:


# Test mtpleasant with the new generic function
testset_mtpleasant = [
    ('106.10-6-67',349557),
    ('106.18-2-61',380530),
    ('112.18-1-32',601769),
]

re_fmv = 'FULL MKT VAL [0-9,]+|FULL MARKET VALUE [0-9,]+'

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mtpleasant'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        print(entry[1])
        print(entry[2])
        raise

failed = []
for key in data['mtpleasant']:
    entry = []
    verbose = False
    for line in data['mtpleasant'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = ext.get_generic(entry,re_fmv)
        output = output.split()
        output = float(output[-1].replace(',',''))
    except:
        failed.append(key)

print(f"Entries without full market value:\n{failed}")


# ### Get taxables

# In[50]:


def get_taxable(entry,taxable_name,verbose=False):
    taxable = None
    taxable_line = find_line(entry,taxable_name)
    if taxable_line == 1:
        #print(f"WARN: {entry[1][0]} doesn't have {taxable_name}")
        return None
    
    taxable = re.search(taxable_name + ' [0-9\\,]+',taxable_line)
    if taxable != None:
        taxable = re.sub(taxable_name + ' ','',taxable.group())
        return float(taxable.replace(',',''))
    else:
        if verbose: print([taxable_line])
        return None
    #else:
    #    taxable = entry[bn][ln].split()
    #    for sn,string in enumerate(taxable):
    #        if string == 'VALUE':
    #            try: taxable = taxable[sn+1]
    #            except: taxable = entry[bn][ln+1]
    #            break
    #if re.search('[a-zA-Z ]', taxable):
    #    taxable = taxable.split()[0]
    #return float(taxable.replace(',',''))


def get_all_taxables(entry,taxable_names,verbose=False):
    taxables = []
    for tn in taxable_names:
        taxables.append(get_taxable(entry,tn,verbose))
    return taxables

# Tests
def run_test(entry,key,result,function,taxable_names):
    output = function(entry,taxable_names)
    assert output == result, f"{key}, {result} != {output}"

testset_bronxville = [
    ('1./1/10',[None,1100000,1100000]),
    ('7.E/3/8',[None,3500000,3500000]),
    ('11./5/1.-311',[None,869261,869261]),
    ('3./3/11',[1475000,1475000,1398580]),
]

testset_cornwall = [
    ('101-1-1',[475000,475000,475000]),
    ('18-3-16',[287400,287400,287400]),
]

testset_scarsdale = [
    ('01.01.1',[1425000,1425000,1425000]),
    ('01.05.12',[1925000,1925000,1925000]),
    ('08.19.32',[800000,800000,800000]),
    ('19.01.40',[1721400,1721400,1721400]),
    ('01.02.20A',[0,0,0]),
    ('11.05.9',[882375,882375,882375]),
]

testset_harrison = [
    ('0011.-1',[2000,2000,2000]),
    ('0883.-31',[12152,12152,11990]),
    ('0103.-13',[1280,1280,1280]),
    ('1005.-32',[19950,19950,19950]),
]

testset_newcastle = [
    ('636.798-760-885',[315,315,315]),
    ('100.11-3-62',[177600,177600,177600]),
    ('80.20-1-10.-82',[38340,38340,24500])
]

taxable_names = ['COUNTY TAXABLE VALUE','VILLAGE TAXABLE VALUE','SCHOOL TAXABLE VALUE']
for key,result in testset_bronxville:
    run_test(data_bronxville[key],key,result,get_all_taxables,taxable_names)

taxable_names = ['COUNTY TAXABLE VALUE','TOWN TAXABLE VALUE','SCHOOL TAXABLE VALUE']
for key,result in testset_cornwall:
    entry = copy.deepcopy(data_cornwall[key][1:])
    run_test(entry,key,result,get_all_taxables,taxable_names)

taxable_names = ['COUNTY TAXABLE','VILLAGE TAXABLE','SCHOOL TAXABLE']
for key,result in testset_scarsdale:
    entry = copy.deepcopy(data_scarsdale[key][1:])
    run_test(entry,key,result,get_all_taxables,taxable_names)

taxable_names = ['COUNTY TAXABLE','TOWN TAXABLE','SCHOOL TAXABLE']
for key,result in testset_harrison:
    entry = copy.deepcopy(data_harrison[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_all_taxables,taxable_names)

taxable_names = ['COUNTY TAXABLE','TOWN TAXABLE','SCHOOL TAXABLE']
for key,result in testset_newcastle:
    entry = copy.deepcopy(data_newcastle[key]['data'])
    entry = unwrap_sublists_recursive(entry)
    entry = break_lines(entry)[1:]
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    run_test(entry,key,result,get_all_taxables,taxable_names)

taxable_names = ['COUNTY TAXABLE VALUE','VILLAGE TAXABLE VALUE','SCHOOL TAXABLE VALUE']
all_taxables = []
for entry in data_bronxville:
    taxable = get_all_taxables(data_bronxville[entry],taxable_names)
    all_taxables.append(taxable)

assert '' in ['', 'test']
assert ['','',''] not in all_taxables
assert None not in all_taxables


# In[51]:


# Test greenburgh
testset_greenburgh = [
    ('6.10-1-10.1',[1487500,1487500,1487500]),
    ('7.280-125-13',[619200,619200,542780]),
    ('8.540-375-12',[979200,979200,979200]),
    ('6.120-103-1.DP1',[None,None,None]),
]

#re_taxables = [
#    'CITY TAXABLE [0-9,]+',
#    'TOWN TAXABLE [0-9,]+',
#    'SCHOOL TAXABLE [0-9,]+'
#]
taxable_names = [
    'CNTY TAXABLE',
    'TOWN TAXABLE',
    'SCHOOL TAXABLE'
]

for key,result in testset_greenburgh:
    entry = []
    for line in data['greenburgh'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = get_all_taxables(entry,taxable_names)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['greenburgh'][key]: print([line])
        print(f"\nEntry:\n{entry}")
        raise


# In[52]:


# Test mamaroneck using the new get_generic function
testset_mamaroneck = [
    ('7-21-103',[1842000,1842000,1760240]),
    ('1-32-400.711',[383400,383400,383400]),
    ('3-13-22',[2200000,2200000,2200000]),
]

taxable_names = [
    '(CNTY TAXABLE|COUNTY TAXABLE VALUE) [0-9,]+',
    '(TOWN TAXABLE( VALUE)?) [0-9,]+',
    '(SCHOOL TAXABLE( VALUE)?) [0-9,]+'
]

for key,result in testset_mamaroneck:
    entry = []
    for line in data['mamaroneck'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mamaroneck'][key]: print([line])
        print("\nEntry:")
        for line in entry: print([line])
        raise

failed = []
for key in data['mamaroneck']:
    entry = []
    for line in data['mamaroneck'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
    except:
        failed.append[key]

print(f"failed to find taxables in entries:\n{failed}")


# In[53]:


# Test bedford using the new get_generic function
testset_bedford = [
    ('49.10-2-4',[61345,61345,39355]),
    ('60.18-1-10',[45525,45525,45525]),
    ('84.15-1-20',[98095,98095,98095]),
]

taxable_names = [
    '(CNTY TAXABLE|COUNTY TAXABLE VALUE) [0-9,]+',
    '(TOWN TAXABLE( VALUE)?) [0-9,]+',
    '(SCHOOL TAXABLE( VALUE)?) [0-9,]+'
]

for key,result in testset_bedford:
    entry = []
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['bedford'][key]: print([line])
        print("\nEntry:")
        for line in entry: print([line])
        raise

failed = []
for key in data['bedford']:
    entry = []
    for line in data['bedford'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
    except:
        failed.append[key]

print(f"failed to find taxables in entries:\n{failed}")


# In[54]:


# Test mtpleasant using the new get_generic function
testset_mtpleasant = [
    ('115.11-5-54',[5200,5200,4150]),
    ('106.15-3-9',[6550,6550,6550]),
    ('112.11-1-4',[10000,10000,10000]),
    ('112.16-1-1',[-44118,-44118,-44118]),
]

taxable_names = [
    '(CNTY TAXABLE|COUNTY TAXABLE( VALUE)?) -?[0-9,]+',
    '(TOWN TAXABLE( VALUE)?) -?[0-9,]+',
    '(SCHOOL TAXABLE( VALUE)?) -?[0-9,]+'
]

for key,result in testset_mtpleasant:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('-?[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
        assert output == result, f"{key}, {result} != {output}"
    except:
        print("Data item:")
        for line in data['mtpleasant'][key]: print([line])
        print("\nEntry:")
        for line in entry: print([line])
        raise

failed = []
for key in data['mtpleasant']:
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        output = []
        for t in taxable_names:
            out = ext.get_generic(entry,t)
            out = re.search('-?[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output.append(out)
    except:
        print(key)
        failed.append[key]

print(f"failed to find taxables in entries:\n{failed}")


# ### Assemble workbook

# In[55]:


wb = Workbook()
ws = wb.active

ws.title = "2024 final roll - Mt Pleasant"
ws['A1'] = 'id'
ws['B1'] = 'OWNERS NAME'
ws['C1'] = 'OWNERS ADDRESS'
ws['D1'] = 'PROPERTY TYPE'
ws['E1'] = 'PROPERTY ADDRESS'
ws['F1'] = 'ZONING'
ws['G1'] = 'ACREAGE'
ws['H1'] = 'FULL MARKET VALUE'
ws['I1'] = 'COUNTY TAXABLE'
ws['J1'] = 'SCHOOL TAXABLE'
ws['K1'] = 'TOWN TAXABLE'


# In[56]:


# Extract the data
taxable_names = {
    'county':'(CNTY TAXABLE|COUNTY TAXABLE( VALUE)?) -?[0-9,]+',
    'town':'(TOWN TAXABLE( VALUE)?) -?[0-9,]+',
    'school':'(SCHOOL TAXABLE( VALUE)?) -?[0-9,]+'
}

row = 2
a = time.time()
failed_acreage = []
for key in data['mtpleasant']:
    print(key)
    ws[f"A{row}"] = key
    
    # Owner names
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:26])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                bool(re.search('EAST [0-9]{6}',entry_line[0])),
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                bool(re.fullmatch('[0-9.]{4} [A-Z]{2} [A-Z]{1}',entry_line[0])),
                bool(re.fullmatch('[0-9]-[0-9A-Z]{2,3}-[0-9]{,2}',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    ws[f"B{row}"] = ', '.join(ext.get_owner_names(entry,key))

    # Owner address
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry_line[0] = str(entry_line[0][:26])
            checks  = [
                'FULL MKT VAL' in entry_line[0],
                'FULL MARKET VALUE' in entry_line[0],
                'DEED BK' in entry_line[0],
                'DEED BOOK' in entry_line[0],
                'EAST-' in entry_line[0],
                bool(re.search('EAST [0-9]{6}',entry_line[0])),
                'ACREAGE' in entry_line[0] or 'ACRES' in entry_line[0],
                'add to' in entry_line[0],
                'add map' in entry_line[0],
                bool(re.search('\bMAP\b',entry_line[0])),
                'TAXABLE' in entry_line[0],
                'CONTIGUOUS PARCEL' in entry_line[0],
                'BANK' in entry_line[0],
                bool(re.search('[A-Z]{2}[0-9]{3}',entry_line[0])),
                bool(re.search('@ [0-9]',entry_line[0])),
                bool(re.fullmatch('#[0-9]+',entry_line[0])),
                bool(re.search('[0-9,]{5,7} EX',entry_line[0])),
                bool(re.fullmatch('[0-9.]{4} [A-Z]{2} [A-Z]{1}',entry_line[0])),
                bool(re.fullmatch('[0-9]-[0-9A-Z]{2,3}-[0-9]{,2}',entry_line[0])),
                'ATTM' in entry_line[0],
                'ATTN:' in entry_line[0],
            ]
            if True in checks:
                continue
            checks = [
                'PRIOR OWNER' in entry_line[0],
            ]
            if True in checks:
                break
            checks = [
                bool(re.match('SUITE',entry_line[0])),
                bool(re.match('UNIT',entry_line[0])),
                bool(re.match('(1ST|2ND|3RD|[0-9]+TH) FLOOR',entry_line[0])),
            ]
            if True in checks:
                entry[-1] = entry[-1] + ' ' + entry_line[0]
            elif 'C/O' in entry_line[0]:
                temp = entry[-1]
                entry[-1] = entry_line[0]
                entry.append(temp)
            else:
                entry.append(entry_line[0])
    ws[f"C{row}"] = ext.get_owner_address(entry,key)

    # Property type
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            for sn, substring in enumerate(entry_line):
                if substring == key and bool(re.fullmatch('[A-Z]{2}',entry_line[sn+1])):
                    del entry_line[sn+1]
                    break
            #print(entry_line)
            entry.append(entry_line)
    ws[f"D{row}"] = get_property_type(entry,key)

    # Property address
    entry = []
    for line in data['mtpleasant'][key]:
        #newline = line.replace('\n','')
        #newline = newline.strip()
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    del entry[0]
    nentry = []
    for line in entry[0]:
        checks = [
            not bool(re.fullmatch('[0-9]+',line)),
            not bool(re.search('ACCT:',line)),
        ]
        if False not in checks:
            nentry.append(line)
    entry[0] = nentry
    for n,e in enumerate(entry):
        entry[n] = list(filter(None,e))
    if len(entry[0]) > 1:
        last_item = entry[0].pop()
        entry[0].insert(0,last_item)
    if len(entry[0]) > 2:
        while len(entry[0]) > 2:
            if 'ACCT' not in entry[0][2]:
                entry[0][1] = entry[0][1] + ' ' + entry[0][2]
                del entry[0][2]
            else:
                break
        entry[0] = entry[0][0:2]
    if len(entry[0]) == 1:
        entry[0].insert(0,'')
    ws[f"E{row}"] = get_property_address(entry,key)

    # Zoning
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    ws[f"F{row}"] = ext.get_generic(entry,re_zoning)

    # Acreage
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    try:
        out = ext.get_generic(entry,re_acreage,verbose)
        out = out.split()
        out = float(out[-1])
    except:
        failed_acreage.append(key)
        out = None
    ws[f"G{row}"] = out

    # Full market value
    entry = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    out = ext.get_generic(entry,re_fmv)
    out = out.split()
    out = float(out[-1].replace(',',''))
    ws[f"H{row}"] = out

    # Taxables
    entry = []
    failed_taxables = []
    for line in data['mtpleasant'][key]:
        if line != '' and line != None and line != []:
            entry_line = re.split('  +',line)
            entry.append(entry_line)
    output = {}
    for t in taxable_names:
        try:
            out = ext.get_generic(entry,taxable_names[t])
            out = re.search('-?[0-9,]+',out).group()
            out = out.replace(',','')
            out = float(out)
            output[t] = out
        except:
            failed_taxables.append((key,t))
            output[t] = None
    ws[f"I{row}"] = output['county']
    ws[f"J{row}"] = output['school']
    ws[f"K{row}"] = output['town']
    row += 1

wb.save('extracted_data.xlsx')
b = time.time()

print(b-a)
print(failed_acreage)

